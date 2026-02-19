import openpyxl as xl #This creates an alias for openpyxl
from openpyxl.chart import BarChart, Reference

def process_workbook (filename):
    # Loading the Excel workbook into a variable that behaves like a dictionary
    # where worksheet names are the key
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # This loop iterates every row stating from row 2, assuming row 1 is a header row,
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=5)

    # Adding a Bar Chart into variable "chart", which was imported in -> "from openpyxl.chart import BarChart, Reference"
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Saves the excel sheet into a new sheet
    wb.save(filename)